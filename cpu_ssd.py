import scrapy
import re
import openpyxl
import datetime
import time
now=datetime.datetime.now()
filename = "CPU_SSD_" + now.strftime("%Y-%m-%d") + ".xlsx"
wb = openpyxl.Workbook()

class coolpcSpider(scrapy.Spider):
	name = "cpu"
	headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36 Edg/140.0.0.0'}
	def start_requests(self):
		urls = [
			'http://www.coolpc.com.tw/evaluate.php',
			'https://browser.geekbench.com/processor-benchmarks/',
			'https://nanoreview.net/en/cpu-list/cinebench-scores',
			'https://nanoreview.net/en/cpu-list/cinebench-scores?page=2'
		]
		yield scrapy.Request(url=urls[0], callback=self.parse)
		yield scrapy.Request(url=urls[1], callback=self.parse_geek)
		yield scrapy.Request(url=urls[2], callback=self.parse_nanoreview, headers=self.headers)
		time.sleep(2)
		yield scrapy.Request(url=urls[3], callback=self.parse_nanoreview2, headers=self.headers)

	def parse(self, response):
		ws = wb.active
		ws.title = "CPU"
		ws.append(["Name", "Cores", "Watt", "Price", "Singel Score", "Multiple Score", "Score/Price"])
		dict_cpu = {}
		for row in response.xpath('//select/optgroup')[3].xpath('.//option'):
			cpu=cores=price=watt=None
			line=row.extract()
			try: print(line)
			except: print(" ==> Can not print"); continue
			price = re.findall(r'\$\d+', line)
			if len(price) > 1: price = price[1]
			elif len(price) == 1: price = price[0]
			else: print(" ==> Does not find price"); continue
			if int(price.strip("$")) < 1000: continue
			line = re.sub(r'(AMD )?R(\d)', r'AMD Ryzen \2', line)
			line = re.sub(r'(AMD )?Athlon', r'AMD Athlon', line)
			line = re.sub('TR', 'Threadripper', line)
			line = re.sub('Intel i', 'Intel Core i', line)
			try: cpu = re.search(r'(Intel|AMD)[\s\-\w\+]+', line).group()
			except AttributeError: continue
			cpu = re.sub(r'[\u4e00-\u9fff]+', '', cpu).strip()
			try: watt = re.search(r'[\/\)]\d+W\/?', line).group().strip("W/").replace("/","").replace(")","")
			except: pass
			try: cores = re.search(r'全?\d+大?核(\/\d+緒)?', line).group()
			except AttributeError: continue
			if dict_cpu.get(cpu) == None: dict_cpu[cpu] = cores, watt, int(price.strip("$"))
			elif dict_cpu.get(cpu)[2] > int(price.strip("$")) and dict_cpu.get(cpu)[2] - int(price.strip("$")) < 5000:
				watt = dict_cpu.get(cpu)[1]
				dict_cpu[cpu] = cores, watt, int(price.strip("$"))
			else: continue
		i=2
		for x, y in dict_cpu.items():
			j=str(i)
			s1 = "=VLOOKUP(A" + j + ",nanoreview!A:B,2,0)"
			s2 = "=VLOOKUP(A" + j + ",nanoreview!A:C,3,0)"
			sp_ratio = "=(E"+j+"*2+F"+j+")/D"+j
			x = re.sub(" MPK", "", x)
			x = re.sub(r'(\d{4})G PRO', r'PRO \1G', x)
			ws.append([ x, y[0], y[1], y[2], s1, s2, sp_ratio ])
			i+=1
		wb.save(filename)
		time.sleep(2)
		
		# SSD
		ws = wb.create_sheet("SSD")
		ws.append(["Name", "Size", "Read", "Write", "Type", "Price", "PCIe Ver", "Warranty", "Form factor"])	
		for row in response.xpath('//select/optgroup')[6].xpath('.//option'):
			name=size=read=write=type=price=pcie=warranty=form=None
			result=row.extract()
			try: print(result)
			except: continue
			try: read = re.search(r'讀:?(\d{3,5})', result).group(1)
			except: continue
			name = re.search(r'^.+?\/', result).group()
			name = re.sub('<.+?>','',name)
			name = name.replace("/", "")
			name = re.sub(r'[^w]:\d+M','',name)
			if "Gen4" in result: pcie="4.0"
			if "P5 Plus" in result: pcie="4.0"
			try: size = re.search(r'\d{3,4}GB?|[^-]\dTB?', result).group()
			except: pass
			if size: 
				size = size.replace(")", "")
				if not "SE20" in name: name = name.replace(size, "")
			else:
				size = re.search(r'\dTB', result).group()
			size = size.replace(" ", "").replace("G", "").replace("B", "")
			if "T" in size:
				size = size.replace("T", "")
				size = int(size)*1024
			name = name.replace('讀:'+read, "")
			write = re.search(r'寫:?(\d{3,5})', result).group(1)
			try: type = re.search('[TQM]LC', result).group()
			except: pass
			list_price = re.findall(r'\$\d+', result)
			if len(list_price) > 1: price = list_price[1]
			elif len(list_price) == 1: price = list_price[0]
			if "三年" in result: warranty = "3 years"
			if "五年" in result: warranty = "5 years"
			if "2.5吋" in result: form = "2.5吋"
			if int(read) > 800: form = "M.2"
			ws.append([name, size, read, write, type, price, pcie, warranty, form])
		wb.save(filename)
		time.sleep(2)
		
	def parse_geek(self, response):
		#wb = openpyxl.load_workbook(filename)
		ws = wb.create_sheet("Geek_single")
		latest = 0
		score = 0
		for row in response.selector.xpath("//tr/td[@class]"):
			line = row.extract()
			if "name" in line:
				cpu = re.search(r'(Intel|AMD|HP|A4)( [\-\w\+]+)+', line).group()
			elif "score" in line: 
				score = re.search(r'\d+', line).group()
			if int(score) - int(latest) > 10000:
				ws = wb.create_sheet("Geek_multiple")
			if int(score) > 0:
				ws.append([cpu, score])
				latest = score
				score = 0
		wb.save(filename)
		time.sleep(2)

	def parse_nanoreview(self, response):
		try: ws = wb["nanoreview"]
		except KeyError: ws = wb.create_sheet("nanoreview")
		cpu = response.xpath('//tr').xpath('./td/div/a/text()').extract()
		score = response.xpath('//div[@style="margin-bottom: 6px;"]/text()').extract()
		single = []
		multiple = []
		for i in range(0, len(score), 4):
			single.append(score[i].strip())
			multiple.append(score[i+1].strip())
		for x, y, z in zip(cpu, single, multiple):
			x = re.sub(r'Core i(\d) ', r'Intel Core i\1-', x)
			x = re.sub("Ryzen", "AMD Ryzen", x)
			x = re.sub("Core Ultra", "Intel Core Ultra", x)
			ws.append([x, y, z])
		wb.save(filename)
		time.sleep(2)
		
	def parse_nanoreview2(self, response):
		try: ws = wb["nanoreview"]
		except KeyError: ws = wb.create_sheet("nanoreview")
		cpu = response.xpath('//tr').xpath('./td/div/a/text()').extract()
		score = response.xpath('//div[@style="margin-bottom: 6px;"]/text()').extract()
		single = []
		multiple = []
		for i in range(0, len(score), 4):
			single.append(score[i].strip())
			multiple.append(score[i+1].strip())
		for x, y, z in zip(cpu, single, multiple):
			x = re.sub(r'Core i(\d) ', r'Intel Core i\1-', x)
			x = re.sub("Ryzen", "AMD Ryzen", x)
			x = re.sub("Core Ultra", "Intel Core Ultra", x)
			ws.append([x, y, z])
		wb.save(filename)
		
