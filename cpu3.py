import scrapy
import re
import openpyxl
import datetime
now=datetime.datetime.now()
filename = "CPU_" + now.strftime("%Y-%m-%d") + ".xlsx"

class coolpcSpider(scrapy.Spider):
	name = "cpu"
	def start_requests(self):
		urls = [
			'http://www.coolpc.com.tw/evaluate.php',
			'https://browser.geekbench.com/processor-benchmarks/'
		]
		yield scrapy.Request(url=urls[0], callback=self.parse)
		yield scrapy.Request(url=urls[1], callback=self.parse_geek)

	def parse(self, response):
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.append(["Name", "Cores", "Watt", "Price", "Singel Score", "Multiple Score", "Score/Price"])
		dict_cpu = {}
		for row in response.xpath('//select/optgroup')[3].xpath('.//option'):
			cpu=cores=price=watt=None
			line=row.extract()
			try: print line
			except: continue
			try: price = re.search('\$\d+', line).group()
			except: continue
			line = re.sub(r'(AMD )?R(\d)', r'AMD Ryzen \2', line)
			line = re.sub('TR', 'Threadripper', line)
			line = re.sub('Intel i', 'Intel Core i', line)
			if "G5400" in line: line=line.replace("G5400", "Gold G5400")
			cpu = re.search('(Intel|AMD)[\s\-\w\+]+', line).group()
			try: watt = re.search('[\/\)]\d+W\/?', line).group().strip("W/").replace("/","").replace(")","")
			except: pass
			if 'Intel Xeon E5-2620 V4' in cpu: watt = "85"
			try: cores = re.search('\d+[^\w\s,]\/\d+[^\w,](GPU)?', line).group()
			except AttributeError: cores = re.search('\d[^\w\s\"]{1,2}(\d[^\w]{1})?', line).group()
			if dict_cpu.get(cpu) == None: dict_cpu[cpu] = cores, watt, int(price.strip("$"))
			elif dict_cpu.get(cpu)[2] > int(price.strip("$")) and dict_cpu.get(cpu)[2] - int(price.strip("$")) < 5000:
				watt = dict_cpu.get(cpu)[1]
				dict_cpu[cpu] = cores, watt, int(price.strip("$"))
			else: continue
		i=2
		for x, y in dict_cpu.items():
			j=str(i)
			s1 = "=VLOOKUP(A" + j + ",Geek_single!A:B,2,0)"
			s2 = "=VLOOKUP(A" + j + ",Geek_multiple!A:B,2,0)"
			sp_ratio = "=(E"+j+"*2+F"+j+")/D"+j
			if "4350G" in x: s1 = 1137; s2 = 4954
			if "4650G" in x: s1 = 1184; s2 = 6994
			if "4750G" in x: s1 = 1239; s2 = 8228
			if "G4930" in x: s1 = 810; s2= 1501
			if "G5900" in x: s1 = 838; s2= 1431
			if "10400F" in x: s1 = 1089; s2 = 6160
			if "10700F" in x: s1 = 1265; s2 = 8090
			if "10700KF" in x: s1 = 1414; s2 = 9098
			ws.append([ x, y[0], y[1], y[2], s1, s2, sp_ratio ])
			i+=1
		wb.save(filename)
		
	def parse_geek(self, response):
		wb = openpyxl.load_workbook(filename)
		ws = wb.create_sheet(title="Geek_single")
		latest = 0
		score = 0
		for row in response.selector.xpath("//tr/td[@class]"):
			line = row.extract()
			#try: print line
			#except: pass
			if "name" in line:
				cpu = re.search('(Intel|AMD|HP|A4)( [\-\w\+]+)+', line).group()
			elif "score" in line: 
				score = re.search('\d+', line).group()
			if int(score) - int(latest) > 10000:
				print "==> change to multiple sheet"
				ws = wb.create_sheet(title="Geek_multiple")
			if score > 0:
				ws.append([cpu, score])
				#print " => cpu: " + cpu
				#print " => score: " + score + "\n"
				latest = score
				score = 0
		wb.save(filename)
